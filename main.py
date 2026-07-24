import os
import json
import requests
import rasterio
import numpy as np
from datetime import datetime
import geopandas as gpd
from rasterstats import zonal_stats
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ======================================================
# URL & PATH CONFIGURATION
# ======================================================
JSON_URL = "https://satda.tmd.go.th/wp-content/uploads/data/json/radar_mosaic_latest.json"

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

PROVINCE_GEOJSON = "province.geojson" 

# ======================================================
# 1. อ่าน JSON & สร้างโฟลเดอร์สำหรับรอบนี้
# ======================================================
print("กำลังดึงข้อมูล JSON...")
response = requests.get(JSON_URL)
response.raise_for_status()
data = response.json()

base_local = data["base_time"]["local"]
dt = datetime.fromisoformat(base_local)
folder_name = dt.strftime("%Y%m%d_%H%M")
save_folder = os.path.join(OUTPUT_DIR, folder_name)
os.makedirs(save_folder, exist_ok=True)

# บันทึก Metadata
with open(os.path.join(save_folder, "metadata.json"), "w", encoding="utf-8") as f:
    json.dump(data, f, indent=4, ensure_ascii=False)

# ======================================================
# 2. ดาวน์โหลด TIFF
# ======================================================
print("กำลังดาวน์โหลดไฟล์ TIFF...")
host = data["host"].rstrip("/")
groups = ["past", "present", "nowcast"]
count = 0

for group in groups:
    for item in data["radar"][group]:
        url = f"{host}{item['path']}/{item['filename']}"
        thai_time = item["time_local"][11:16].replace(":", "")
        new_filename = item["filename"].replace(".tiff", f"_{thai_time}.tiff")
        outfile = os.path.join(save_folder, new_filename)

        r = requests.get(url, stream=True)
        if r.status_code == 200:
            with open(outfile, "wb") as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            count += 1

print(f"ดาวน์โหลดสำเร็จทั้งหมด {count} ไฟล์")

# ======================================================
# 3. ประมวลผล Forecast TIFF -> Rain Accumulation
# ======================================================
print("กำลังประมวลผลฝนสะสม...")
tiff_files = sorted([
    os.path.join(save_folder, f) 
    for f in os.listdir(save_folder) if f.endswith(".tiff")
])

if len(tiff_files) < 17:
    raise Exception("จำนวน TIFF ไม่ครบ 17 ไฟล์")

forecast_files = tiff_files[5:17]

# อ่านไฟล์แรกตั้งต้น
with rasterio.open(forecast_files[0]) as src:
    meta = src.meta.copy()
    nodata = src.nodata
    dbz = src.read(1).astype(np.float32)

valid = (dbz != nodata)
Z = np.power(10.0, dbz / 10.0)
R = np.power(Z / 200.0, 1.0 / 1.6)
accum = np.where(valid, R * 0.25, 0.0)

# รวมอีก 11 ไฟล์
for file in forecast_files[1:]:
    with rasterio.open(file) as src:
        dbz = src.read(1).astype(np.float32)
        valid = (dbz != nodata)
        Z = np.power(10.0, dbz / 10.0)
        R = np.power(Z / 200.0, 1.0 / 1.6)
        rain = np.where(valid, R * 0.25, 0.0)
        accum += rain

accum = accum.astype(np.float32)
accum[~np.isfinite(accum)] = nodata

# Save Accumulated TIFF
meta.update(dtype="float32", nodata=nodata)
output_file = os.path.join(save_folder, "Forecast_Accumulated_Rainfall_3hr_mm.tif")
with rasterio.open(output_file, "w", **meta) as dst:
    dst.write(accum, 1)

# ======================================================
# 4. Zonal Statistics & Export Excel
# ======================================================
print("กำลังคำนวณ Zonal Statistics และสร้าง Excel...")
gdf = gpd.read_file(PROVINCE_GEOJSON)
stats = zonal_stats(gdf, output_file, stats=["max", "mean", "min"], nodata=nodata)

result = pd.DataFrame(stats)
result["จังหวัด"] = gdf["PROV_NAMT"]
result = result[["จังหวัด", "max", "mean", "min"]]
result.columns = ["จังหวัด", "MAX_mm", "MEAN_mm", "MIN_mm"]

excel_file = os.path.join(save_folder, "Province_Rainfall_MAX_MEAN_MIN.xlsx")
result = result.sort_values(by="MAX_mm", ascending=False)
result.to_excel(excel_file, index=False)

# แต่งฟอร์แมต Excel
wb = load_workbook(excel_file)
ws = wb.active
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"

for row in ws.iter_rows(min_row=2):
    row[1].number_format = "0.0"
    row[2].number_format = "0.0"
    row[3].number_format = "0.0"

for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 4

wb.save(excel_file)
print("สร้างไฟล์ Excel เรียบร้อย:", excel_file)

# ======================================================
# 5. ส่งไฟล์และรายงานเข้า Discord
# ======================================================
DISCORD_WEBHOOK_URL = os.environ.get("DISCORD_WEBHOOK_URL")

if DISCORD_WEBHOOK_URL:
    print("กำลังส่งรายงานเข้า Discord...")
    top3 = result.head(3)
    summary_text = f"🌧️ **รายงานคาดการณ์ฝนสะสม 3 ชั่วโมง (รอบ {dt.strftime('%d/%m/%Y %H:%M')})**\n"
    summary_text += "🏆 **3 อันดับจังหวัดฝนตกหนักสุด:**\n"
    for idx, row in top3.iterrows():
        summary_text += f"- **{row['จังหวัด']}**: Max {row['MAX_mm']:.1f} mm (เฉลี่ย {row['MEAN_mm']:.1f} mm)\n"
    summary_text += "\n📁 *แนบไฟล์รายละเอียด Excel ด้านล่างนี้ครับ*"

    with open(excel_file, "rb") as f:
        payload = {"content": summary_text}
        files = {"file": (f"Province_Rainfall_{folder_name}.xlsx", f)}
        r = requests.post(DISCORD_WEBHOOK_URL, data=payload, files=files)
        
    if r.status_code in [200, 204]:
        print("ส่งรายงานเข้า Discord สำเร็จ!")
    else:
        print(f"เกิดข้อผิดพลาดในการส่ง Discord: {r.status_code} {r.text}")
else:
    print("ไม่พบ DISCORD_WEBHOOK_URL ข้ามขั้นตอนการส่ง Discord")
